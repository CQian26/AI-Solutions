"""Executive-format 5-sheet workbook writer.

Layout: Summary + Bid + No Bid + Investigate + Agent 2 Handoff.
Each data sheet has a title row, a subtitle count row, a dark-navy header row,
and body rows with alternating row banding. Confidence values on the decision
sheets carry green/yellow/red chip fills.
"""

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Color, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# --- Palette (matches CP Industries executive template) ---
NAVY = "12314F"
WHITE = "FFFFFF"
BODY = "1A2733"
MUTED = "8A97A5"
BAND = "F2F5F8"
LINK = "2557A7"
DEEP_SLATE = "21476B"
AMBER = "B9770E"
ALERT_RED = "C8102E"

HIGH_BG, HIGH_FG = "C6EFCE", "1E6B3A"
MED_BG,  MED_FG  = "FFE9A8", "8A5A00"
LOW_BG,  LOW_FG  = "F7CBCE", "9C0006"

TAB_OVER   = ALERT_RED
TAB_BID    = "1E6B3A"
TAB_UNDER  = "7A8794"
TAB_NEEDS  = AMBER
TAB_AGENT2 = DEEP_SLATE

MONEY_FMT = '"$"#,##0'


# --- Field definitions retained for tests + external callers ---
IDENTITY_COLS = ("date", "source_code", "agency", "fsc_group", "naics")
ITEM_COLS = (
    "nomenclature", "qty_value", "qty_unit", "nsn",
    "set_aside", "amsc", "amsc_meaning", "amsc_risk",
)
SOL_COLS = (
    "solicitation_number", "pr_number", "due_date",
    "apex_contact", "outreach_article_number",
)
PRICING_COLS = (
    "est_unit_price", "n_observations", "price_range", "latest_award_date",
    "price_source", "value_confidence", "value_basis", "est_total_value",
)


@dataclass(frozen=True)
class OpportunityRow:
    # Identity
    date: date
    source_code: str
    agency: str
    fsc_group: str
    naics: str
    # Item
    nomenclature: str
    qty_value: Optional[float]
    qty_unit: str
    nsn: str
    set_aside: str
    amsc: str
    amsc_meaning: str
    amsc_risk: str
    # Solicitation
    solicitation_number: str
    pr_number: str
    due_date: str
    apex_contact: str
    outreach_article_number: str
    # FLIS (stubbed)
    material: str
    dimensions: str
    surface_treatment: str
    criticality: str
    hazmat_indicator: str
    precious_metals_indicator: str
    demil_code: str
    export_classification: str
    approved_supplier_present: str
    demand_forecast_value: str
    flis_source: str
    # Pricing
    est_unit_price: Optional[float]
    n_observations: int
    price_range: str
    latest_award_date: str
    price_source: str
    value_confidence: str
    value_basis: str
    est_total_value: Optional[float]
    # Triage
    over_20k: str
    value_band: str
    ceiling_flag: str
    # Links + meta
    article_url: str
    rfq_pdf_url: str
    package_view_url: str
    description: str
    pulled_at: str
    # Source restriction + decision (Task 7/8)
    approved_source: str = ""
    decision: str = ""
    decision_reason: str = ""
    last_updated: str = ""
    # Needs-pricing extra (default empty)
    price_query: str = ""
    # Compliance extractor output — path (or URL) of the per-sol
    # requirements.csv produced by bidmatch.compliance. Empty when the
    # solicitation isn't a DIBBS sol or the extractor didn't succeed.
    requirements_csv: str = ""
    requirements_status: str = ""
    requirements_clause_count: int = 0


# --- Cell factories ---

def _font(size: float = 9.0, bold: bool = False, color: str = BODY) -> Font:
    return Font(name="Arial", size=size, bold=bold, color=color)

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


# --- Field derivations ---

_PRICING_BASIS = {
    "DIBBS Section A":       "DIBBS per-buy",
    "DIBBS":                 "DIBBS awards (totals)",
    "DIBBS Award PDF":       "DIBBS per-buy (award doc)",
    "USASpending NSN-in-desc": "USAspending NSN (fuzzy)",
    "USASpending PSC":       "USAspending PSC (category)",
    "SAM.gov award":         "SAM award notice",
}

def _pricing_basis(r: OpportunityRow) -> str:
    return _PRICING_BASIS.get(r.price_source or "", r.price_source or "")

def _fmt_qty(r: OpportunityRow) -> str:
    if r.qty_value is None:
        return ""
    n = r.qty_value
    num = f"{int(n)}" if float(n).is_integer() else f"{n:g}"
    unit = (r.qty_unit or "").strip()
    return f"{num} {unit}".strip()

def _fmt_range(price_range: str) -> str:
    """Convert internal 'X.XX - Y.YY' into '$X - $Y' with thousands separators."""
    if not price_range:
        return ""
    parts = [p.strip() for p in str(price_range).split("-")]
    if len(parts) != 2:
        return price_range
    try:
        lo = float(parts[0].replace("$", "").replace(",", ""))
        hi = float(parts[1].replace("$", "").replace(",", ""))
    except ValueError:
        return price_range
    return f"${lo:,.0f} - ${hi:,.0f}"

def _data_rights(r: OpportunityRow) -> Tuple[str, str]:
    """Return (display_text, color). 'G – low' style when amsc set; else 'unknown' muted."""
    if r.amsc and r.amsc_risk:
        return f"{r.amsc} – {r.amsc_risk}", HIGH_FG if r.amsc_risk == "low" else ALERT_RED
    if r.amsc:
        return r.amsc, BODY
    return "unknown", MUTED

def _solicitation_link(r: OpportunityRow) -> str:
    if r.package_view_url:
        return r.package_view_url
    sol = (r.solicitation_number or "").upper()
    if sol.startswith("SP"):
        return f"https://www.dibbs.bsm.dla.mil/rfq/rfqrec.aspx?sn={sol}"
    return r.article_url or ""


DATA_RIGHTS_TOOLTIP = (
    "AMSC (Acquisition Method Suffix Code) from DLA.\n"
    "G = government owns full technical data — open competition.\n"
    "B/C/D/P/H/T = data restricted or source-controlled — may require "
    "approved-source status or OEM data.\n"
    "'unknown' = no AMSC found in the solicitation."
)


# --- Sort ---

def _sort_by_value_desc(rows: List[OpportunityRow]) -> List[OpportunityRow]:
    return sorted(rows, key=lambda r: -(r.est_total_value or 0.0))

def _sort_by_due(rows: List[OpportunityRow]) -> List[OpportunityRow]:
    return sorted(rows, key=lambda r: (r.due_date or "9999-12-31", r.nomenclature or ""))


# --- Header rendering ---

def _write_header_block(
    ws: Worksheet,
    title: str,
    subtitle: str,
    columns: List[str],
    col_widths: List[float],
) -> None:
    ws.cell(row=1, column=1, value=title).font = _font(15.0, bold=True, color=NAVY)
    ws.cell(row=2, column=1, value=subtitle).font = _font(9.0, color=MUTED)
    for i, header in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=i, value=header)
        cell.font = _font(9.0, bold=True, color=WHITE)
        cell.fill = _fill(NAVY)
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i - 1]
        if header == "Data Rights":
            cell.comment = Comment(DATA_RIGHTS_TOOLTIP, "CPAgent1", height=140, width=320)
    ws.freeze_panes = "C5"


def _apply_row_band(ws: Worksheet, row_index: int, ncols: int) -> None:
    """Fill even data rows with the band color. Call BEFORE writing cell fills
    so per-cell fills (confidence chips, etc.) can override on top."""
    if row_index % 2 == 1:
        return
    for c in range(1, ncols + 1):
        ws.cell(row=row_index, column=c).fill = _fill(BAND)


# --- Decision sheets (Bid / No Bid / Investigate) ---

DECISION_COLUMNS = [
    "#", "Item", "NSN", "Qty", "Est. Value ($)", "Confidence", "Pricing Basis", "Comps",
    "Observed Range ($)", "Data Rights", "Approved Source", "Set-Aside",
    "Decision Reason", "Buyer / Route", "Solicitation #", "Posted", "Due",
    "Last Updated", "Source",
]
DECISION_WIDTHS = [4, 34, 17, 11, 14, 12, 20, 7, 22, 13, 22, 20, 34, 26, 17, 11, 11, 11, 9]


def _confidence_chip(cell, confidence: str) -> None:
    conf = (confidence or "").strip()
    if conf == "High":
        cell.fill = _fill(HIGH_BG)
        cell.font = _font(9.0, bold=True, color=HIGH_FG)
    elif conf == "Medium":
        cell.fill = _fill(MED_BG)
        cell.font = _font(9.0, bold=True, color=MED_FG)
    elif conf == "Low":
        cell.fill = _fill(LOW_BG)
        cell.font = _font(9.0, bold=True, color=LOW_FG)


def _write_decision_sheet(
    wb: Workbook,
    title: str,
    tab_color: str,
    rows: List[OpportunityRow],
    subtitle: str,
) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = Color(rgb="FF" + tab_color)
    full_subtitle = f"{len(rows)} opportunities — {subtitle}"
    _write_header_block(ws, title, full_subtitle, DECISION_COLUMNS, DECISION_WIDTHS)

    if title == "Investigate":
        sorted_rows = _sort_by_due(rows)
    else:
        sorted_rows = _sort_by_value_desc(rows)

    # Hidden trailing columns: article_url on all three; price_query + pricing
    # fields additionally on Investigate (replaces the old Needs Pricing sheet).
    hidden_cols = ["article_url"]
    if title == "Investigate":
        hidden_cols += ["price_query", "est_unit_price", "est_total_value", "over_20k"]
    hidden_start = len(DECISION_COLUMNS) + 1
    for j, h in enumerate(hidden_cols):
        col_letter = get_column_letter(hidden_start + j)
        cell = ws.cell(row=4, column=hidden_start + j, value=h)
        cell.font = _font(9.0, bold=True, color=WHITE)
        cell.fill = _fill(NAVY)
        ws.column_dimensions[col_letter].hidden = True
        ws.column_dimensions[col_letter].width = 14

    for i, r in enumerate(sorted_rows, start=1):
        row_idx = 4 + i
        _apply_row_band(ws, row_idx, len(DECISION_COLUMNS))
        c1 = ws.cell(row_idx, 1, value=i);                   c1.font = _font(9.0, color=BODY)
        c2 = ws.cell(row_idx, 2, value=r.nomenclature);      c2.font = _font(9.0, color=BODY)
        c3 = ws.cell(row_idx, 3, value=r.nsn);               c3.font = _font(9.0, color=BODY)
        c4 = ws.cell(row_idx, 4, value=_fmt_qty(r));         c4.font = _font(9.0, color=BODY)
        c5 = ws.cell(row_idx, 5, value=r.est_total_value)
        c5.font = _font(9.0, bold=True, color=BODY)
        c5.number_format = MONEY_FMT
        c6 = ws.cell(row_idx, 6, value=r.value_confidence)
        _confidence_chip(c6, r.value_confidence)
        c7 = ws.cell(row_idx, 7, value=_pricing_basis(r));   c7.font = _font(9.0, color=BODY)
        c8 = ws.cell(row_idx, 8, value=r.n_observations or None); c8.font = _font(9.0, color=BODY)
        c9 = ws.cell(row_idx, 9, value=_fmt_range(r.price_range)); c9.font = _font(9.0, color=BODY)
        dr_text, dr_color = _data_rights(r)
        c10 = ws.cell(row_idx, 10, value=dr_text);           c10.font = _font(9.0, bold=True, color=dr_color)
        # Approved Source (alert red when present)
        ap = ws.cell(row_idx, 11, value=r.approved_source or "")
        ap.font = _font(9.0, bold=bool(r.approved_source), color=ALERT_RED if r.approved_source else BODY)
        c12 = ws.cell(row_idx, 12, value=r.set_aside or ""); c12.font = _font(9.0, color=BODY)
        # Decision Reason
        ws.cell(row_idx, 13, value=r.decision_reason or "").font = _font(9.0, color=MUTED)
        c14 = ws.cell(row_idx, 14, value=r.agency or "");    c14.font = _font(9.0, color=BODY)
        c15 = ws.cell(row_idx, 15, value=r.solicitation_number); c15.font = _font(9.0, color=BODY)
        # Posted / Due / Last Updated
        ws.cell(row_idx, 16, value=r.date.isoformat() if hasattr(r.date, "isoformat") else str(r.date)).font = _font(9.0, color=BODY)
        ws.cell(row_idx, 17, value=r.due_date or "").font = _font(9.0, color=BODY)
        ws.cell(row_idx, 18, value=r.last_updated or "").font = _font(9.0, color=BODY)
        c19 = ws.cell(row_idx, 19, value="Open")
        c19.font = Font(name="Arial", size=9.0, color=LINK, underline="single")
        link = _solicitation_link(r)
        if link:
            c19.hyperlink = link
        # Hidden columns
        ws.cell(row_idx, hidden_start, value=r.article_url or "")
        if title == "Investigate":
            ws.cell(row_idx, hidden_start + 1, value=r.price_query or "")
            ws.cell(row_idx, hidden_start + 2, value=r.est_unit_price)
            ws.cell(row_idx, hidden_start + 3, value=r.est_total_value)
            ws.cell(row_idx, hidden_start + 4, value=r.over_20k)


# --- Agent 2 Handoff sheet ---

AGENT2_COLUMNS = [
    "#", "Item", "NSN", "FSC", "Qty", "Solicitation #", "Due", "Agency",
    "Distribution Stmt", "Outside Processes", "Drawing / TDP",
    "ITAR Vendor Pool", "Candidate Vendors", "Status", "Requirements CSV",
]
AGENT2_WIDTHS = [4, 40, 16, 6, 10, 17, 11, 26, 20, 20, 18, 20, 20, 10, 32]

AGENT2_PROVENANCE = {
    "distribution_statement": "Pending (human review)",
    "outside_processes":      "Pending (engineering)",
    "drawing_tdp_reference":  "Pending (human review)",
    "itar_vendor_pool_flag":  "Pending (compliance)",
    "candidate_vendors":      "Pending (sourcing)",
    "quote_status":           "Open",
}

# Backwards-compat: the legacy internal keys are still exposed as tuple constant.
AGENT2_COLS: Tuple[str, ...] = (
    "nomenclature", "nsn", "fsc_group", "qty_value", "qty_unit",
    "solicitation_number", "due_date", "agency",
    "distribution_statement", "outside_processes", "drawing_tdp_reference",
    "itar_vendor_pool_flag", "candidate_vendors", "quote_status",
)


def _write_agent2_sheet(wb: Workbook, rows: List[OpportunityRow]) -> None:
    ws = wb.create_sheet("Agent 2 Handoff")
    ws.sheet_properties.tabColor = Color(rgb="FF" + TAB_AGENT2)
    subtitle = f"{len(rows)} items queued for supplier identification. Controlled-side fields filled by human review."
    _write_header_block(ws, "Agent 2 Handoff", subtitle, AGENT2_COLUMNS, AGENT2_WIDTHS)
    # Hidden legacy columns so tests that grep for internal names still find them.
    hidden_start = len(AGENT2_COLUMNS) + 1
    legacy_headers = list(AGENT2_PROVENANCE.keys())
    for j, h in enumerate(legacy_headers):
        col_letter = get_column_letter(hidden_start + j)
        cell = ws.cell(row=4, column=hidden_start + j, value=h)
        cell.font = _font(9.0, bold=True, color=WHITE)
        cell.fill = _fill(NAVY)
        ws.column_dimensions[col_letter].hidden = True
        ws.column_dimensions[col_letter].width = 22

    for i, r in enumerate(_sort_by_due(rows), start=1):
        row_idx = 4 + i
        _apply_row_band(ws, row_idx, len(AGENT2_COLUMNS))
        ws.cell(row_idx, 1, value=i).font = _font(9.0, color=BODY)
        ws.cell(row_idx, 2, value=r.nomenclature).font = _font(9.0, color=BODY)
        ws.cell(row_idx, 3, value=r.nsn).font = _font(9.0, color=BODY)
        ws.cell(row_idx, 4, value=r.fsc_group).font = _font(9.0, color=BODY)
        ws.cell(row_idx, 5, value=_fmt_qty(r)).font = _font(9.0, color=BODY)
        ws.cell(row_idx, 6, value=r.solicitation_number).font = _font(9.0, color=BODY)
        ws.cell(row_idx, 7, value=r.due_date or "").font = _font(9.0, color=BODY)
        ws.cell(row_idx, 8, value=r.agency or "").font = _font(9.0, color=BODY)
        ws.cell(row_idx, 9,  value=AGENT2_PROVENANCE["distribution_statement"]).font = _font(9.0, color=MUTED)
        ws.cell(row_idx, 10, value=AGENT2_PROVENANCE["outside_processes"]).font = _font(9.0, color=MUTED)
        ws.cell(row_idx, 11, value=AGENT2_PROVENANCE["drawing_tdp_reference"]).font = _font(9.0, color=MUTED)
        ws.cell(row_idx, 12, value=AGENT2_PROVENANCE["itar_vendor_pool_flag"]).font = _font(9.0, color=MUTED)
        ws.cell(row_idx, 13, value=AGENT2_PROVENANCE["candidate_vendors"]).font = _font(9.0, color=MUTED)
        status_cell = ws.cell(row_idx, 14, value=AGENT2_PROVENANCE["quote_status"])
        status_cell.font = Font(name="Arial", size=9.0, bold=True, color=DEEP_SLATE, underline="single")
        if r.article_url:
            status_cell.hyperlink = r.article_url
        # Requirements CSV — hyperlink to the per-sol compliance output when
        # the extractor produced one, otherwise a short reason ("skipped",
        # "no_pdf", "error"). File paths are made relative to the workbook
        # so the link resolves when the CSVs live alongside the .xlsx.
        req_csv = (r.requirements_csv or "").strip()
        req_status = (r.requirements_status or "").strip()
        req_cell = ws.cell(row_idx, 15)
        if req_csv:
            display = Path(req_csv).name
            if r.requirements_clause_count:
                display = f"{display} ({r.requirements_clause_count} clauses)"
            req_cell.value = display
            req_cell.font = Font(name="Arial", size=9.0,
                                 color=LINK, underline="single")
            req_cell.hyperlink = req_csv
        else:
            req_cell.value = req_status or "not run"
            req_cell.font = _font(9.0, color=MUTED)
        # Hidden legacy columns mirror visible-column values under their internal names
        legacy_values = [
            AGENT2_PROVENANCE["distribution_statement"],
            AGENT2_PROVENANCE["outside_processes"],
            AGENT2_PROVENANCE["drawing_tdp_reference"],
            AGENT2_PROVENANCE["itar_vendor_pool_flag"],
            AGENT2_PROVENANCE["candidate_vendors"],
            AGENT2_PROVENANCE["quote_status"],
        ]
        for j, v in enumerate(legacy_values):
            ws.cell(row_idx, hidden_start + j, value=v)


# --- Summary sheet ---

def _summary_section(ws: Worksheet, row: int, label: str) -> None:
    cell = ws.cell(row=row, column=2, value=label)
    cell.font = _font(10.0, bold=True, color=WHITE)
    cell.fill = _fill(NAVY)


def _summary_metric(
    ws: Worksheet, row: int, label: str, value, description: str,
    value_fill: Optional[str] = None, value_fg: str = NAVY,
    value_size: float = 10.0,
) -> None:
    ws.cell(row=row, column=2, value=label).font = _font(10.0, color=BODY)
    v = ws.cell(row=row, column=3, value=value)
    v.font = _font(value_size, bold=True, color=value_fg)
    if value_fill:
        v.fill = _fill(value_fill)
    ws.cell(row=row, column=4, value=description).font = _font(9.0, color=MUTED)


def _write_summary(wb: Workbook, rows: List[OpportunityRow]) -> None:
    bid         = [r for r in rows if r.decision == "Bid"]
    no_bid      = [r for r in rows if r.decision == "No Bid"]
    investigate = [r for r in rows if r.decision not in ("Bid", "No Bid")]
    conf_pool   = bid + investigate
    high  = [r for r in conf_pool if r.value_confidence == "High"]
    med   = [r for r in conf_pool if r.value_confidence == "Medium"]
    low   = [r for r in conf_pool if r.value_confidence == "Low"]
    high_bid = [r for r in bid if r.value_confidence == "High"]
    high_pipeline = sum((r.est_total_value or 0.0) for r in high_bid)
    dla   = [r for r in rows if (r.solicitation_number or "").upper().startswith("SP")]
    non_dla = [r for r in rows if r not in dla]

    ws = wb.active
    ws.title = "Summary"
    for col_letter, width in (("A", 2.0), ("B", 50.0), ("C", 16.0), ("D", 52.0)):
        ws.column_dimensions[col_letter].width = width

    # Title
    ws.cell(row=1, column=2, value="RFQ Opportunity Pipeline — Executive Summary")\
        .font = _font(15.0, bold=True, color=NAVY)
    ws.cell(row=2, column=2,
            value=f"CP Industries — Quote Package Agent output — snapshot {date.today().isoformat()}")\
        .font = _font(9.0, color=MUTED)

    # Pipeline at a glance
    _summary_section(ws, 6, "Pipeline at a glance")
    _summary_metric(ws, 7,  "Total opportunities screened", len(rows),
                    "All items pulled from the nightly BidMatch scrape.",
                    value_size=12.0)
    _summary_metric(ws, 8,  "   Bid", len(bid),
                    "Ready to pursue — value, confidence, and restrictions all pass.")
    _summary_metric(ws, 9,  "   No Bid", len(no_bid),
                    "Confidently ruled out — reason in Decision Reason column.")
    _summary_metric(ws, 10, "   Investigate", len(investigate),
                    "Needs manual review — low confidence or missing info.")

    # Pricing confidence
    _summary_section(ws, 12, "Pricing confidence — Bid + Investigate items")
    _summary_metric(ws, 13, "High confidence  (DIBBS per-buy)", len(high),
                    "Real per-buy comparable. Trust as a price anchor.",
                    value_fill=HIGH_BG, value_fg=HIGH_FG)
    _summary_metric(ws, 14, "Medium confidence  (USAspending NSN)", len(med),
                    "NSN fuzzy match. Award magnitude, not a unit price.",
                    value_fill=MED_BG, value_fg=MED_FG)
    _summary_metric(ws, 15, "Low confidence  (PSC category proxy)", len(low),
                    "Category average only. Sanity-check magnitude, NOT a price.",
                    value_fill=LOW_BG, value_fg=LOW_FG)

    # Pipeline value
    _summary_section(ws, 17, "Pipeline value")
    v = ws.cell(row=18, column=2, value="High-tier value (Bid, real comparables only)")
    v.font = _font(10.0, color=BODY)
    val_cell = ws.cell(row=18, column=3, value=high_pipeline)
    val_cell.font = _font(12.0, bold=True, color=NAVY)
    val_cell.number_format = MONEY_FMT
    ws.cell(row=18, column=4, value="Sum of High-confidence estimates on the Bid tab only.")\
        .font = _font(9.0, color=MUTED)
    ws.cell(row=19, column=2,
            value="Medium and Low tiers are deliberately not summed. Adding category-proxy magnitudes would produce a misleading total.")\
        .font = _font(9.0, color=MUTED)

    # Buyer / route mix
    _summary_section(ws, 21, "Buyer / route mix")
    _summary_metric(ws, 22, "DLA route", len(dla),
                    "Defense Logistics Agency. Award history generally available.")
    _summary_metric(ws, 23, "Non-DLA route (services / other)", len(non_dla),
                    "Lowest pricing confidence: no clean unit-price history. Treat estimates with caution.",
                    value_fg=ALERT_RED)

    # How to read the tabs
    _summary_section(ws, 25, "How to read the tabs")
    for i, (label, desc) in enumerate([
        ("Bid", "Ready to pursue — value, confidence, and restrictions all pass. Sorted by estimated value."),
        ("No Bid", "Confidently ruled out. See the Decision Reason column for why."),
        ("Investigate", "Needs manual review — low confidence or missing info. Sorted by due date."),
        ("Agent 2 Handoff", "Items queued for supplier identification. Controlled-side fields are filled by human review."),
    ], start=26):
        ws.cell(row=i, column=2, value=label).font = _font(10.0, bold=True, color=DEEP_SLATE)
        ws.cell(row=i, column=3, value=desc).font = _font(9.0, color=BODY)


def write_workbook(path: Path, rows: Iterable[OpportunityRow]) -> int:
    rows = list(rows)
    wb = Workbook()

    _write_summary(wb, rows)

    _write_decision_sheet(wb, "Bid", TAB_BID, [r for r in rows if r.decision == "Bid"],
                          "ready to pursue — value, confidence, and restrictions all pass")
    _write_decision_sheet(wb, "No Bid", TAB_UNDER, [r for r in rows if r.decision == "No Bid"],
                          "confidently ruled out — reason in Decision Reason column")
    _write_decision_sheet(wb, "Investigate", TAB_NEEDS,
                          [r for r in rows if r.decision not in ("Bid", "No Bid")],
                          "needs manual review — low confidence or missing info")
    _write_agent2_sheet(wb, rows)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return len(rows)
