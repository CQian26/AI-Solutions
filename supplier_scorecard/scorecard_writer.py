#!/usr/bin/env python3
"""
scorecard_writer.py
===================
Assemble a supplier-scorecard workbook from a list of per-contract results.

Emits an .xlsx with THREE sheets:

  1. "Scorecard"    — one row per unique FAR/DFARS clause across all contracts:
       citation, regulation, part, title, count, coverage%, contract IDs
       Sorted by count (desc). This is the primary "what codes matter" view.

  2. "Contract x Clause" — pivot matrix: rows = clauses, columns = contracts,
       cell = "✓" if that clause appears in that contract, else blank.
       Column headers freeze; first column freezes.

  3. "Raw Contracts" — one row per contract, showing metadata pulled from
       SAM.gov + which sources we scanned (description, N attachments) +
       count of clauses found.

Consumes the shape produced by run.py's pipeline:
    {
      "contracts": [
         {
           "id":        "SPE7L026T0325",        # solicitation# (or notice ID)
           "title":     "...",
           "agency":    "...",
           "naics":     "336413",
           "psc":       "2590",
           "set_aside": "SBA",
           "posted":    "2026-06-05",
           "deadline":  "2026-06-22",
           "url":       "https://sam.gov/opp/...",
           "sources":   ["description", "attachment:sow.pdf", ...],
           "clauses":   [ {"regulation":"FAR","number":"52.212-4","title":"..."}, ... ],
           "notes":     "...",
         },
         ...
      ]
    }
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def apply_filters(agg_rows: list[dict], filters_cfg: dict) -> tuple[list[dict], list[dict]]:
    """Split aggregated clause rows into (kept, dropped) based on filters_cfg.

    Each dropped row gains a 'drop_reason' field explaining which filter caught it.
    """
    kept: list[dict] = []
    dropped: list[dict] = []

    boiler = filters_cfg.get("drop_admin_boilerplate") or {}
    boiler_on = boiler.get("enabled")
    boiler_set = set(boiler.get("citations") or [])
    boiler_reason = boiler.get("reason") or "Admin boilerplate"

    regs = filters_cfg.get("drop_regulations") or {}
    regs_on = regs.get("enabled")
    regs_set = set(regs.get("regulations") or [])
    regs_reason = regs.get("reason") or "Regulation excluded"

    below = filters_cfg.get("drop_below_count") or {}
    below_on = below.get("enabled")
    below_min = int(below.get("min_contracts") or 2)
    below_reason = below.get("reason") or f"Appeared in fewer than {below_min} contracts"

    specific = filters_cfg.get("drop_specific") or {}
    specific_on = specific.get("enabled")
    specific_set = set(specific.get("citations") or [])
    specific_reason = specific.get("reason") or "Manually excluded"

    for r in agg_rows:
        reasons = []
        if boiler_on and r["citation"] in boiler_set:
            reasons.append(boiler_reason)
        if regs_on and r["regulation"] in regs_set:
            reasons.append(regs_reason)
        if below_on and r["count"] < below_min:
            reasons.append(below_reason)
        if specific_on and r["citation"] in specific_set:
            reasons.append(specific_reason)

        if reasons:
            r = {**r, "drop_reason": " | ".join(reasons)}
            dropped.append(r)
        else:
            kept.append(r)
    return kept, dropped


def _load_filters(path: Optional[Path]) -> dict:
    if not path:
        return {}
    p = Path(path)
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ODD_FILL = PatternFill("solid", fgColor="F2F2F2")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = ws["A2"]


def _autosize(ws, min_w: int = 8, max_w: int = 60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        best = 0
        for cell in col:
            v = cell.value
            if v is None:
                continue
            best = max(best, min(max_w, max(len(str(v).split("\n")[0]), min_w)))
        ws.column_dimensions[letter].width = best + 2


def write_scorecard(results: dict, out_path: Path, *, filters_path: Optional[Path] = None) -> Path:
    contracts = results.get("contracts", []) or []
    filters_cfg = _load_filters(filters_path)
    # Detect simulation mode by looking at the notes column.
    is_simulated = any(
        (c.get("notes") or "").upper().startswith(("MOCK", "DEMO MOCK", "SIMULAT"))
        for c in contracts
    )

    # --- aggregate ---
    agg: dict[str, dict] = {}
    for c in contracts:
        cid = c.get("id") or c.get("title") or "?"
        for cl in c.get("clauses") or []:
            key = f"{cl['regulation']} {cl['number']}"
            slot = agg.setdefault(key, {
                "regulation": cl["regulation"],
                "number": cl["number"],
                "citation": key,
                "part": cl["number"].rsplit("-", 1)[0],
                "title": cl.get("title") or "",
                "explanation": cl.get("explanation") or "",
                "contracts": set(),
                "titles_seen": set(),
            })
            slot["contracts"].add(cid)
            if cl.get("title"):
                slot["titles_seen"].add(cl["title"])
            if not slot["explanation"] and cl.get("explanation"):
                slot["explanation"] = cl["explanation"]

    total_contracts = len(contracts) or 1
    rows = []
    for key, s in agg.items():
        rows.append({
            "citation": s["citation"],
            "regulation": s["regulation"],
            "part": s["part"],
            "number": s["number"],
            "title": s["title"] or (next(iter(s["titles_seen"])) if s["titles_seen"] else ""),
            "explanation": s["explanation"],
            "count": len(s["contracts"]),
            "coverage": len(s["contracts"]) / total_contracts,
            "contract_ids": sorted(s["contracts"]),
        })
    rows.sort(key=lambda r: (-r["count"], r["regulation"], r["number"]))

    # Split into kept / dropped per the filters config (if any).
    kept_rows, dropped_rows = apply_filters(rows, filters_cfg) if filters_cfg else (rows, [])
    rows = kept_rows  # main sheets show only kept clauses

    wb = Workbook()

    # === Simulation banner (only appears when the run used mock data) =======
    if is_simulated:
        ws_warn = wb.active
        ws_warn.title = "!! SIMULATED DATA !!"
        ws_warn.sheet_properties.tabColor = "C00000"
        banner_fill = PatternFill("solid", fgColor="C00000")
        banner_font = Font(color="FFFFFF", bold=True, size=14)
        lines = [
            "⚠ THIS SCORECARD IS SIMULATED — DO NOT USE FOR BID/NO-BID DECISIONS",
            "",
            "Every clause and contract row was fabricated from FSC × agency templates",
            "because the pipeline could not reach api.sam.gov (or was explicitly told",
            "to use --dev-only-mock-dir). The clause distribution is plausible but is",
            "NOT the actual clause set for these opportunities.",
            "",
            "To get a real scorecard:",
            "  1. Get a free key at https://open.gsa.gov/api/get-opportunities-public-api/",
            "  2. export SAM_API_KEY=your_key",
            "  3. Re-run the pipeline in a network environment that can reach sam.gov.",
            "",
            "Run this workbook again after those steps and this sheet will disappear.",
        ]
        for i, line in enumerate(lines, start=1):
            c = ws_warn.cell(row=i, column=1, value=line)
            c.font = banner_font
            c.alignment = LEFT
            if i == 1:
                c.fill = banner_fill
        ws_warn.column_dimensions["A"].width = 90

    # === Sheet 1: Scorecard ============================================
    # Columns: Citation | Part | Title | What it means | Count | Contract IDs
    if is_simulated:
        ws = wb.create_sheet("Scorecard")
    else:
        ws = wb.active
        ws.title = "Scorecard"
    ws.append(["Citation", "Part", "Title", "What it means", "Count", "Contract IDs"])
    for i, r in enumerate(rows, start=2):
        ws.append([
            r["citation"], r["part"], r["title"],
            r["explanation"] or "",
            r["count"],
            ", ".join(r["contract_ids"]),
        ])
        # zebra
        if i % 2 == 0:
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = ODD_FILL
        ws.cell(row=i, column=3).alignment = LEFT  # Title
        ws.cell(row=i, column=4).alignment = LEFT  # Explanation
        ws.cell(row=i, column=5).alignment = CENTER  # Count
        ws.cell(row=i, column=6).alignment = LEFT  # Contract IDs
    _style_header(ws, 6)
    _autosize(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 70   # Explanation gets the most room
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 45

    # === Sheet 2: Contract x Clause matrix =============================
    ws2 = wb.create_sheet("Contract x Clause")
    contract_ids = [c.get("id") or c.get("title") or f"C{i+1}"
                    for i, c in enumerate(contracts)]
    header = ["Citation", "Title"] + contract_ids
    ws2.append(header)
    for i, r in enumerate(rows, start=2):
        row = [r["citation"], r["title"]]
        for cid in contract_ids:
            row.append("✓" if cid in r["contract_ids"] else "")
        ws2.append(row)
        if i % 2 == 0:
            for c in range(1, len(header) + 1):
                ws2.cell(row=i, column=c).fill = ODD_FILL
        # ticks centered
        for c in range(3, len(header) + 1):
            ws2.cell(row=i, column=c).alignment = CENTER
    _style_header(ws2, len(header))
    _autosize(ws2)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 55
    for c_idx in range(3, len(header) + 1):
        ws2.column_dimensions[get_column_letter(c_idx)].width = 16
    ws2.freeze_panes = "C2"

    # === Sheet 3: MIL-Specs / Standards ================================
    #
    # Technical standards (MIL-STD, MIL-PRF, MIL-DTL, FED-STD, A-A-...) live
    # here — separate from FAR/DFARS clauses because they're different
    # compliance surfaces (regs vs. technical requirements).
    std_agg: dict[str, dict] = {}
    for c in contracts:
        cid = c.get("id") or "?"
        for s in c.get("standards") or []:
            key = s["citation"]
            slot = std_agg.setdefault(key, {
                "citation": key, "kind": s["kind"], "number": s["number"],
                "contracts": set(),
            })
            slot["contracts"].add(cid)
    std_rows = sorted(
        std_agg.values(),
        key=lambda r: (-len(r["contracts"]), r["kind"], r["number"]),
    )
    ws_std = wb.create_sheet("MIL-Specs & Standards")
    ws_std.append(["Citation", "Type", "Number", "Count", "Coverage %", "Contract IDs"])
    for i, r in enumerate(std_rows, start=2):
        ws_std.append([
            r["citation"], r["kind"], r["number"], len(r["contracts"]),
            len(r["contracts"]) / total_contracts,
            ", ".join(sorted(r["contracts"])),
        ])
        if i % 2 == 0:
            for c in range(1, 7):
                ws_std.cell(row=i, column=c).fill = ODD_FILL
        ws_std.cell(row=i, column=4).alignment = CENTER
        cov = ws_std.cell(row=i, column=5); cov.number_format = "0.0%"; cov.alignment = CENTER
        ws_std.cell(row=i, column=6).alignment = LEFT
    _style_header(ws_std, 6)
    _autosize(ws_std)
    ws_std.column_dimensions["F"].width = 45

    # === Sheet 4: Raw Contracts ========================================
    ws3 = wb.create_sheet("Raw Contracts")
    ws3.append([
        "Contract ID", "Title", "Agency", "NAICS", "PSC/FSC", "Set-Aside",
        "Posted", "Deadline", "Sources scanned", "# Clauses", "# Standards",
        "URL", "Notes",
    ])
    for i, c in enumerate(contracts, start=2):
        ws3.append([
            c.get("id"), c.get("title"), c.get("agency"), c.get("naics"),
            c.get("psc"), c.get("set_aside"), c.get("posted"), c.get("deadline"),
            ", ".join(c.get("sources") or []),
            len(c.get("clauses") or []),
            len(c.get("standards") or []),
            c.get("url"), c.get("notes"),
        ])
        if i % 2 == 0:
            for col in range(1, 14):
                ws3.cell(row=i, column=col).fill = ODD_FILL
        ws3.cell(row=i, column=2).alignment = LEFT
        ws3.cell(row=i, column=9).alignment = LEFT
    _style_header(ws3, 13)
    _autosize(ws3)
    ws3.column_dimensions["B"].width = 45
    ws3.column_dimensions["C"].width = 25
    ws3.column_dimensions["I"].width = 40
    ws3.column_dimensions["L"].width = 45

    # === Sheet 5: Dropped Clauses (audit trail) ========================
    # Same shape as Scorecard + Drop reason column
    if dropped_rows:
        ws4 = wb.create_sheet("Dropped Clauses")
        ws4.append([
            "Citation", "Part", "Title", "What it means",
            "Count", "Contract IDs", "Drop reason",
        ])
        dropped_rows.sort(key=lambda r: (-r["count"], r["regulation"], r["number"]))
        for i, r in enumerate(dropped_rows, start=2):
            ws4.append([
                r["citation"], r["part"], r["title"],
                r.get("explanation") or "",
                r["count"],
                ", ".join(r["contract_ids"]),
                r.get("drop_reason", ""),
            ])
            if i % 2 == 0:
                for col in range(1, 8):
                    ws4.cell(row=i, column=col).fill = ODD_FILL
            ws4.cell(row=i, column=3).alignment = LEFT
            ws4.cell(row=i, column=4).alignment = LEFT
            ws4.cell(row=i, column=5).alignment = CENTER
            ws4.cell(row=i, column=6).alignment = LEFT
            ws4.cell(row=i, column=7).alignment = LEFT
        _style_header(ws4, 7)
        _autosize(ws4)
        ws4.column_dimensions["A"].width = 22
        ws4.column_dimensions["B"].width = 10
        ws4.column_dimensions["C"].width = 45
        ws4.column_dimensions["D"].width = 70
        ws4.column_dimensions["E"].width = 8
        ws4.column_dimensions["F"].width = 45
        ws4.column_dimensions["G"].width = 60
        # Grey out the tab to signal these are excluded.
        ws4.sheet_properties.tabColor = "888888"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
